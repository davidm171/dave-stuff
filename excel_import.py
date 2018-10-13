
def import_map(lang):
    import openpyxl


    #wb = openpyxl.load_workbook('L:\\docs\\documentation\\g099\\01-manage\\06-Work Instructions\\17 - help files\\guide to flare project map.xlsx')
    wb = openpyxl.load_workbook('C:\\spreadsheet\\guide to flare project map.xlsx')
    
    sheetName = "Dir to Guide Mapping - " +  lang.upper()
    print "The sheet name is ", sheetName
    try:
        guideMapSheet = wb[sheetName]            ## Should put check in here so it knows what sheets are there, and ignores if they're not
    except KeyError:
        x = raw_input("The worksheet ", sheetName, " does not exist. Press return to continue without creating cross refs.")
                
     
    # for i in range(1,15, 1):
        # print i, guideMapSheet.cell(row = i, column = 1).value, guideMapSheet.cell(row = i, column = 2).value
        
    guideMap = {}

    for row in range(1, guideMapSheet.max_row):
        if (guideMapSheet['A' + str(row)].value != None) and (guideMapSheet['B' + str(row)].value != None):
            guideName = guideMapSheet['A' + str(row)].value
            mapValue = guideMapSheet['B' + str(row)].value
            print guideName, " <-----> ", mapValue, "\n"

        guideMap[guideName] = mapValue
        
    #print guideMap
    
    return guideMap
